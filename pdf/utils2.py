
import io
from calendar import monthrange
from datetime import date, datetime, timedelta

import openpyxl
from dateutil.relativedelta import relativedelta
from django.http import FileResponse
from reportlab.lib import colors, utils
from reportlab.lib.pagesizes import landscape, letter
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from .models import Beneficios_Mala, Folha_de_Ponto, Funcionario


def importar_excel_beneficios(arquivo):
    workbook = openpyxl.load_workbook(arquivo)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        Beneficios_Mala.objects.update_or_create(
            id=row[0],
            defaults={
                'comp': row[1],
                'codigo': row[2],
                'codigo_fc': row[3],
                'aut': row[4],
                'data_inicio': row[5],
                'data_fim': row[6],
                'dias_calculados': row[7],
                'tipo_de_beneficio': row[8],
                'valor_pago': row[9],
                'data_de_pagamento': row[10],
            }
        )



def importar_excel_folha(arquivo):
    workbook = openpyxl.load_workbook(arquivo)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        Folha_de_Ponto.objects.update_or_create(
            id=row[0],
            defaults={
                'comp': row[1],
                'codigo': row[2],
                'codigo_fc': row[3],
                'total_de_horas': row[4],
                'hs_normais': row[5],
                'hs_noturnas': row[6],
                'he_50': row[7],
                'he_100': row[8],
                '21_m': row[9],
                '21_h': row[10],
                '21_d': row[11],
                '22_m': row[12],
                '22_h': row[13],
                '22_d': row[14],
                '23_m': row[15],
                '23_h': row[16],
                '23_d': row[17],
                '24_m': row[18],
                '24_h': row[19],
                '24_d': row[20],
                '25_m': row[21],
                '25_h': row[22],
                '25_d': row[23],
                '26_m': row[24],
                '26_h': row[25],
                '26_d': row[26],
                '27_m': row[27],
                '27_h': row[28],
                '27_d': row[29],
                '28_m': row[30],
                '28_h': row[31],
                '28_d': row[32],
                '29_m': row[33],
                '29_h': row[34],
                '29_d': row[35],
                '30_m': row[36],
                '30_h': row[37],
                '30_d': row[38],
                '31_m': row[39],
                '31_h': row[40],
                '31_d': row[41],
                '1_m_s': row[42],
                '1_h_s': row[43],
                '1_d_s': row[44],
                '2_m_s': row[45],
                '2_h_s': row[46],
                '2_d_s': row[47],
                '3_m_s': row[48],
                '3_h_s': row[49],
                '3_d_s': row[50],
                '4_m_s': row[51],
                '4_h_s': row[52],
                '4_d_s': row[53],
                '5_m_s': row[54],
                '5_h_s': row[55],
                '5_d_s': row[56],
                '6_m_s': row[57],
                '6_h_s': row[58],
                '6_d_s': row[59],
                '7_m_s': row[60],
                '7_h_s': row[61],
                '7_d_s': row[62],
                '8_m_s': row[63],
                '8_h_s': row[64],
                '8_d_s': row[65],
                '9_m_s': row[66],
                '9_h_s': row[67],
                '9_d_s': row[68],
                '10_m_s': row[69],
                '10_h_s': row[70],
                '10_d_s': row[71],
                '11_m_s': row[72],
                '11_h_s': row[73],
                '11_d_s': row[74],
                '12_m_s': row[75],
                '12_h_s': row[76],
                '12_d_s': row[77],
                '13_m_s': row[78],
                '13_h_s': row[79],
                '13_d_s': row[80],
                '14_m_s': row[81],
                '14_h_s': row[82],
                '14_d_s': row[83],
                '15_m_s': row[84],
                '15_h_s': row[85],
                '15_d_s': row[86],
                '16_m_s': row[87],
                '16_h_s': row[88],
                '16_d_s': row[89],
                '17_m_s': row[90],
                '17_h_s': row[91],
                '17_d_s': row[92],
                '18_m_s': row[93],
                '18_h_s': row[94],
                '18_d_s': row[95],
                '19_m_s': row[96],
                '19_h_s': row[97],
                '19_d_s': row[98],
                '20_m_s': row[99],
                '20_h_s': row[100],
                '20_d_s': row[101],
                '21_m_s': row[102],
                '21_h_s': row[103],
                '21_d_s': row[104],
                '22_m_s': row[105],
                '22_h_s': row[106],
                '22_d_s': row[107],
                '23_m_s': row[108],
                '23_h_s': row[109],
                '23_d_s': row[110],
                '24_m_s': row[111],
                '24_h_s': row[112],
                '24_d_s': row[113],
                '25_m_s': row[114],
                '25_h_s': row[115],
                '25_d_s': row[116],
                '26_m_s': row[117],
                '26_h_s': row[118],
                '26_d_s': row[119],
                '27_m_s': row[120],
                '27_h_s': row[121],
                '27_d_s': row[122],
                '28_m_s': row[123],
                '28_h_s': row[124],
                '28_d_s': row[125],
                '29_m_s': row[126],
                '29_h_s': row[127],
                '29_d_s': row[128],
                '30_m_s': row[129],
                '30_h_s': row[130],
                '30_d_s': row[131],
                '31_m_s': row[132],
                '31_h_s': row[133],
                '31_d_s': row[134],
                'periodo_apontamento': row[135],
            
            }
        )

def competencia_meses_adjacentes(competencia: int) -> list:
    try:
        competencia_str = str(competencia)
        data = datetime.strptime(competencia_str, "%Y%m")
    except ValueError:
        return "Formato de competência inválido. Deve ser um inteiro no formato AAAAMM."

    # Subtrai um mês para obter a competência anterior
    mes_anterior = data - timedelta(days=1)
    mes_anterior_int = int(mes_anterior.strftime("%Y%m"))

    # Adiciona um mês para obter a competência seguinte
    mes_seguinte = data + timedelta(days=31)
    mes_seguinte_int = int(mes_seguinte.strftime("%Y%m"))

    return [mes_anterior_int, competencia, mes_seguinte_int]


def competencianormal(competencia: int) -> list:
    try:
        competencia_str = str(competencia)
        data = datetime.strptime(competencia_str, "%Y%m")
    except ValueError:
        return "Formato de competência inválido. Deve ser um inteiro no formato AAAAMM."

    # Subtrai um mês para obter a competência anterior
    mes_anterior = data - timedelta(days=1)
    mes_anterior_int = int(mes_anterior.strftime("%Y%m"))

    # Adiciona um mês para obter a competência seguinte
    mes_seguinte = data + timedelta(days=31)
    mes_seguinte_int = int(mes_seguinte.strftime("%Y%m"))

    return [competencia]




def get_image(path, width):
    image = utils.ImageReader(path)
    aspect_ratio = float(image.getSize()[1]) / float(image.getSize()[0])
    height = aspect_ratio * width
    return path, width, height


def draw_centered_text(p, y, text, fontsize=11, fontname="Helvetica", fontstyle="normal"):
    if fontstyle == "bold":
        fontname = f"{fontname}-Bold"

    p.setFont(fontname, fontsize)
    text_width = p.stringWidth(text, fontname, fontsize)
    x = (letter[0] - text_width) / 2.2
    p.drawString(x, y, text)

def draw_info_rect(p, x, y, width, height, text, fontsize=11, fontname="Helvetica", fontstyle="normal"):
    p.rect(x, y, width, height)
    if fontstyle == "bold":
        fontname = f"{fontname}-Bold"
    p.setFont(fontname, fontsize)
    text_width = p.stringWidth(text, fontname, fontsize)
    text_x = x + (width - text_width) / 2
    text_y = y + height / 2 - fontsize / 2
    p.drawString(text_x, text_y, text)


def gerar_pdf2(funcionario):
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer)


    # Configurar o título do arquivo PDF
    p.setTitle(f"{funcionario.comp} - {funcionario.codigo_fc} - {funcionario.nome}.pdf")


    logo_path, logo_width, logo_height = get_image('static/images/go2b3.jpg', 138)
    p.drawImage(logo_path, 210, 780, width=logo_width, height=logo_height)
    # Desenhe as informações do funcionário no PDF
    draw_centered_text(p, 750, f"EXTRATO SIMPLES - POR COLABORADOR – FATO GERADOR", fontsize=10, fontstyle="bold")
    draw_centered_text(p, 700, f"COMPETÊNCIA: {funcionario.comp}", fontsize=10, fontstyle="bold")
    draw_centered_text(p, 675, f"Contrato: {funcionario.cto}")
    draw_centered_text(p, 650, f"Matrícula: {funcionario.codigo_fc}")
    draw_centered_text(p, 625, f"Nome: {funcionario.nome}")
    draw_centered_text(p, 600, f"Cargo: {funcionario.cargo}")
    draw_centered_text(p, 575, f"Admissão: {funcionario.adm}")
    draw_centered_text(p, 550, f"Demissão: {funcionario.dem_compt}")
    draw_centered_text(p, 525, f"CPF: {funcionario.cpf}")
    draw_centered_text(p, 500, f"Cliente: {funcionario.cliente}")


    draw_centered_text(p, 425, f"IDENTIFICAÇÃO COMPROVANTE/EVIDÊNCIAS:", fontsize=10, fontstyle="bold")
    draw_centered_text(p, 400, f"• AUSÊNCIAS LEGAIS/FÉRIAS/DECIMO TERC./VERBAS RESCISÓRIAS/SAL.MATERN: VIDE AUTENTICAÇÕES E RECIBO.", fontsize=10, fontstyle="bold")
    draw_centered_text(p, 375, f"• FGTS RESCISÓRIO E FGTS SOBRE ACIDENTE TRABALHO: VIDE SEFIP E GRRF.", fontsize=10, fontstyle="bold")





    table_widthxx = 550

    rect_xx = (letter[0] - table_widthxx) / 2

    p.setFont("Helvetica", 10)
    p.drawString(rect_xx, 273, f"_________________________________________________________________________________________________")
    p.drawString(rect_xx, 181, f"_________________________________________________________________________________________________")
    
    
    p.setFont("Helvetica", 7)
    p.drawString(rect_xx + 275, 150, f"-------")
    p.drawString(rect_xx + 325, 150, f"-------")
    p.drawString(rect_xx + 400, 150, f"-------")
    p.drawString(rect_xx + 275, 135, f"-------")
    p.drawString(rect_xx + 325, 135, f"-------")
    p.drawString(rect_xx + 400, 135, f"-------")
    p.drawString(rect_xx + 275, 120, f"-------")
    p.drawString(rect_xx + 325, 120, f"-------")
    p.drawString(rect_xx + 400, 120, f"-------")
    p.drawString(rect_xx + 275, 105, f"-------")
    p.drawString(rect_xx + 325, 105, f"-------")
    p.drawString(rect_xx + 400, 105, f"-------")
    p.drawString(rect_xx + 275, 90, f"-------")
    p.drawString(rect_xx + 325, 90, f"-------")
    p.drawString(rect_xx + 400, 90, f"-------")

    draw_centered_text(p, 325, f"COMPROVANTE PAGAMENTO - COMPETÊNCIA: {funcionario.comp}", fontsize=12, fontstyle="bold")
    draw_centered_text(p, 300, f"FOLHA/RCT", fontsize=12, fontstyle="bold")

    p.setFont("Helvetica-Bold", 10)
    p.drawString(rect_xx + 10, 275, f"Dados Consultados")
    p.drawString(rect_xx + 10, 166, f"Autenticação")
    p.drawString(rect_xx + 200, 166, f"Data")
    p.drawString(rect_xx + 275, 166, f"Banco")
    p.drawString(rect_xx + 325, 166, f"Agência")
    p.drawString(rect_xx + 400, 166, f"Conta")
    p.drawString(rect_xx + 475, 166, f"Valor R$")


    p.setFont("Helvetica", 8)
    p.drawString(rect_xx + 10, 250, f"Agência:")
    p.drawString(rect_xx + 10, 235, f"Conta:")
    p.drawString(rect_xx + 10, 220, f"Descrição Lote:")
    p.drawString(rect_xx + 10, 205, f"Situação Lote:")
    p.drawString(rect_xx + 10, 190, f"Favorecidos:")
    p.drawString(rect_xx + 10 + 80, 250, f"1195-9 (BANCO DO BRASIL) OU 3380-4 (BRADESCO)")
    p.drawString(rect_xx + 10 + 80, 235, f"106742-7 (BANCO DO BRASIL) OU 15801-1 (BRADESCO)")
    p.drawString(rect_xx + 10 + 80, 220, f"PAG DIVERS DOC – CREDITO CONTA SALÁRIO")
    p.drawString(rect_xx + 10 + 80, 205, f"PROCESSADO - EFETUADO")


    p.setFont("Helvetica", 7)
    p.drawString(rect_xx + 10 + 80, 190, f"{funcionario.codigo_fc} - {funcionario.nome}")
    p.drawString(rect_xx + 10, 150, f"{funcionario.aut_1}")
    p.drawString(rect_xx + 200, 150, f"{funcionario.data_1}")
    p.drawString(rect_xx + 475, 150, f"{funcionario.liquido_1}")
    p.drawString(rect_xx + 10, 135, f"{funcionario.aut_2}")
    p.drawString(rect_xx + 200, 135, f"{funcionario.data_2}")
    p.drawString(rect_xx + 475, 135, f"{funcionario.liquido_2}")
    p.drawString(rect_xx + 10, 120, f"{funcionario.aut_3}")
    p.drawString(rect_xx + 200, 120, f"{funcionario.data_3}")
    p.drawString(rect_xx + 475, 120, f"{funcionario.liquido_3}")
    p.drawString(rect_xx + 10, 105, f"{funcionario.aut_4}")
    p.drawString(rect_xx + 200, 105, f"{funcionario.data_4}")
    p.drawString(rect_xx + 475, 105, f"{funcionario.liquido_4}")
    p.drawString(rect_xx + 10, 90, f"{funcionario.aut_5}")
    p.drawString(rect_xx + 200, 90, f"{funcionario.data_5}")
    p.drawString(rect_xx + 475, 90, f"{funcionario.liquido_5}")













    # Finalizar a primeira página e iniciar a segunda página
    p.showPage()



    # Desenhar as informações na segunda página
    draw_centered_text(p, 805, f"Recibo de Pagamento", fontsize=16, fontstyle="bold")
    p.setFont("Helvetica", 6) # Altere o segundo argumento para o tamanho desejado
    p.drawString(465, 787, f"Sofware Responsável http://www.gi.com.br") 
    
    # Defina a largura da tabela
    table_width = 550
    # Desenhar um retângulo com informações do funcionário acima da tabela
    rect_x = (letter[0] - table_width) / 2
    # Aqui mexe na altura de onde fica na pagina.
    rect_y = -24
    # Altura do retangulo
    rect_height = 60
    p.roundRect(rect_x, rect_y + 750, table_width, rect_height, 1, stroke=1, fill=0)
    p.setFont("Helvetica", 8)
    p.drawString(rect_x + 10, rect_y + 750 + 48, f"Código:")
    p.drawString(rect_x + 10, rect_y + 750 + 35, f"{funcionario.codigo_fc}")
    p.drawString(rect_x + 50, rect_y + 750 + 48, f"Nome do Funcionário:")
    p.drawString(rect_x + 50, rect_y + 750 + 35, f"{funcionario.nome}")
    p.drawString(rect_x + 210, rect_y + 750 + 48, f"Função:")
    p.drawString(rect_x + 210, rect_y + 750 + 35, f"{funcionario.cargo}")
    p.drawString(rect_x + 345, rect_y + 750 + 48, f"Admissão:")
    p.drawString(rect_x + 345, rect_y + 750 + 35, f"{funcionario.adm}")
    p.drawString(rect_x + 400, rect_y + 750 + 48, f"Demissão:")
    p.drawString(rect_x + 400, rect_y + 750 + 35, f"{funcionario.dem_compt}")
    p.drawString(rect_x + 480, rect_y + 750 + 48, f"Competência:")
    p.drawString(rect_x + 480, rect_y + 750 + 35, f"{funcionario.comp}")


    data = [
        ['Cód. Descrição', 'Referência', 'Vencimentos', 'Descontos'],
        ['HORAS NORMAIS', f"{funcionario.qtde_hs_normais}", f"{funcionario.horas_normais}", ' '],
        ['D.S.R. S/HORAS NORMAL', ' ', f"{funcionario.dsr_s_horas_normal}", ' '],
        ['D.S.R. FERIADO', f"{funcionario.qtde_dsr_feriado}", f"{funcionario.dsr_feriado}", ' '],
        ['HORA EXTRA 100% / HORA EXTRA 100% NOT', f"{funcionario.qtde_he_100} / {funcionario.qtde_he_100_not}", f"{funcionario.hora_extra_100} / {funcionario.hora_extra_100_noturno}", ' '],
        ['D.S.R. S/HORA EXTRA 100%', ' ', f"{funcionario.dsr_s_hora_extra_100}", ' '],
        ['HORA EXTRA 50% / HORA EXTRA 50% NOT', f"{funcionario.qtde_he_50} / {funcionario.qtde_he_50_not}", f"{funcionario.hora_extra_50} / {funcionario.hora_extra_50_noturno}", ' '],
        ['D.S.R. S/HORA EXTRA 50%', f" ", f"{funcionario.dsr_s_hora_extra_50}", ' '],
        ['ADIC. PERICULOSIDADE', ' ', f"{funcionario.adic_periculosidade}", ' '],
        ['ADICIONAL NOTURNO', ' ', f"{funcionario.adicional_noturno}", ' '],
        ['D.S.R. S/ADICIONAL', ' ', f"{funcionario.dsr_s_adicional}", ' '],
        ['ADICIONAL DE FUNÇÃO 25%', ' ', f"{funcionario.adicional_de_funcao_25}", ' '],
        ['SALARIO FAMILIA', ' ', f"{funcionario.salario_familia}", ' '],
        ['FALTA ABONADA-PONTO ELETR.', ' ', f"{funcionario.falta_abonada_efeito_visualizacao}", ' '],
        ['LICENÇA GESTANTE (LEI 14.151)', ' ', f"{funcionario.licenca_remunerada_gestante}", ' '],
        ['ATESTADO HORISTAS', ' ', f"{funcionario.atestado_horistas}", ' '],
        ['SAL. MATERNIDADE', ' ', f"{funcionario.salario_maternidade}", ' '],
        ['AUX. DOENÇA / ACID. TRABALHO (15 DIAS)', ' ', f"{funcionario.aux_doenca_15_dias} / {funcionario.acidente_de_trabalho_15_dias}", ' '],
        ['VERBAS RESCISÓRIAS (Art 7º CF)', ' ', f"{funcionario.verbas_rescisorias}", ' '],
        ['FERIAS', ' ', f"{funcionario.ferias}", ' '],
        ['1/3 FERIAS', ' ', f"{funcionario.um_terco_ferias}", ' '],
        ['13º SALARIO INDENIZADO E ADICIONAIS', ' ', f"{funcionario.decimo_terceiro_salario_indenizado_e_adicionais_considerar}", ' '],
        ['ARREDONDAMENTO', ' ', f"{funcionario.arredondamento}", ' '],
        ['REEMBOLSO EXAME MEDICO/EPI/UNIF', ' ', f"{funcionario.dev_desc_exame_medico_epi_unif}", ' '],
        ['DIF. VR / VA  - DIF. VALE TRANSPORTE', ' ', f"{funcionario.dif_vale_refeicao} - {funcionario.dif_vale_transporte}", ' '],
        ['SALDO NEGATIVO', ' ', f"{funcionario.saldo_negativo_verba_nao_repassada}", ' '],
        ['DESC. FALTAS (DIAS+ATRASOS) E HORAS IND.', f"{funcionario.qtde_dias_e_hs_desconto}", ' ', f"{funcionario.desc_faltas_dias_atrasos_e_horas_indevidas}"],
        ['DESC. D.S.R. S/FALTAS (DIAS)', ' ', ' ', f"{funcionario.desc_dsr_s_faltas_dias}"],
        ['FALTAS ABONADAS', ' ', ' ', f"{funcionario.falta_abonada_efeito_visualizacao}"],
        ['DESC. ARREDONDAMENTO', ' ', ' ', f"{funcionario.desc_arredondamento}"],
        ['DESC. AVISO', ' ', ' ', f"{funcionario.desc_aviso}"],
        ['DESC. I.N.S.S./DESC. I.R.R.F', ' ', ' ', f"{funcionario.desc_inss_desc_irrf}"],
        ['DESC. I.N.S.S. S/13º SALARIO – INSS (Férias)', ' ', ' ', f"{funcionario.desc_inss_s_13_salario}"],
        ['SEGURO VIDA', ' ', f" ", f"{funcionario.seguro_vida}"],
        ['DESC. ASSIST. ODONTOLOGICA', ' ', ' ', f"{funcionario.desc_assist_odontologica}"],
        ['DESC. VALE REFEICAO NAO UTILIZADO', ' ', ' ', f"{funcionario.desc_vale_refeicao_nao_utilizado}"],
        ['DESC. VR/VA', ' ', ' ', f"{funcionario.desc_vr_va}"],
        ['DESC UNIFORME / EPI', ' ', ' ', f"{funcionario.desc_uniforme_epi_div_sind_judic_adiant_mater}"],
        ['DESC. VALE-TRANSPORTE NAO UTILIZADO', ' ', ' ', f"{funcionario.desc_vale_transporte_nao_utilizado}"],
        ['DESC. VALE-TRANSPORTE', ' ', ' ', f"{funcionario.desc_vale_transporte}"],
        ['DESC. SALDO NEGATIVO', ' ', ' ', f"{funcionario.desc_saldo_negativo}"],
        [' ', ' ', ' ', ' '],
        [' ', ' ', ' ', ' '],
        [' ', ' ', ' ', ' '],
        [' ', ' ', ' ', ' '],
        [' ', ' ', ' ', ' '],



    ]



    table = Table(data)

    # Definir a largura das colunas
    table._argW[0], table._argW[1], table._argW[2], table._argW[3] = 265, 95, 95, 95

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.white),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),  #Espaçemnto da primeira linha.
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ('LINEABOVE', (0, 1), (-1, 1), 1, colors.black),  # Adicionado
    ]))


    table.wrapOn(p, letter[0], letter[1])
    table_x = (letter[0] - table_width) / 2
    table_y = rect_y - rect_height + 120  # Subtrair 2 para diminuir ainda mais o espaço entre a tabela e o retângulo

    table.drawOn(p, table_x, table_y)  #


########################################################################################################
#  QUADRO FINAL DO CONTRA CHEQUE.
########################################################################################################

#################################
# QUADRADO GRANDE
#################################

    # Defina a largura da tabela
    table_width = 550
    # Desenhar um retângulo com informações do funcionário acima da tabela
    rect_x = (letter[0] - table_width) / 2
    # Aqui mexe na altura de onde fica na pagina.
    rect_y = -714
    # Altura do retangulo
    rect_height = 90
    p.roundRect(rect_x, rect_y + 750, table_width, rect_height, 1, stroke=1, fill=0)

#################################
# QUADRADO DE BAIXO PARA ASSINATURA
#################################

    table_width2 = 550
    # Desenhar um retângulo com informações do funcionário acima da tabela
    rect_x2 = (letter[0] - table_width) / 2
    # Aqui mexe na altura de onde fica na pagina.
    rect_y2 = -714
    # Altura do retangulo
    rect_height2 = 40
    p.roundRect(rect_x2, rect_y2 + 750, table_width2, rect_height2, 1, stroke=1, fill=0)

#################################
# QUADRADO GRANDE PARA A DIREITA
#################################

    table_width3 = 275
    # Mexe onde ele fica visando lado
    rect_x3 = 306
    # Aqui mexe na altura de onde fica na pagina.
    rect_y3 = -714
    # Altura do retangulo
    rect_height3 = 90
    p.roundRect(rect_x3, rect_y3 + 750, table_width3, rect_height3, 0, stroke=1, fill=0)

#################################
# LINHA NA DIREITA GRANDE PARA A DIREITA
#################################


#################################
# LINHA NA DIREITA GRANDE PARA A DIREITA
#################################

    table_width3 = 137.5
    # Mexe onde ele fica visando lado
    rect_x3 = 306
    # Aqui mexe na altura de onde fica na pagina.
    rect_y3 = -674
    # Altura do retangulo
    rect_height3 = 50
    p.roundRect(rect_x3, rect_y3 + 750, table_width3, rect_height3, 0, stroke=1, fill=0)

########################################################################################################
    p.drawString(rect_x + 300 + 10, rect_y + 750 + 70, f"Total Vencimentos:")
    p.drawString(rect_x + 300 + 20, rect_y + 750 + 55, f"R$ {funcionario.vencimentos}")
    p.drawString(rect_x + 445 + 10, rect_y + 750 + 70, f"Total Descontos:")
    p.drawString(rect_x + 445 + 20, rect_y + 750 + 55, f"R$ {funcionario.descontos}")
    p.drawString(rect_x + 350, rect_y + 750 + 20, f"Valor Líquido ==== R$ {funcionario.liquido}")
    p.drawString(rect_x + 25, rect_y + 750 + 22, f"__________________________________")
    p.drawString(rect_x + 25, rect_y + 750 + 10, f"Declaro ter recebido a importância líquida discriminada neste recibo")

    p.setFont("Helvetica", 10)
    p.drawString(rect_x + 25, rect_y + 750 + 25, f"{funcionario.nome}")

    # Finalizar a segunda página e começa a 3°.
    p.showPage()




    table_widthxx = 550

    rect_xx = (letter[0] - table_widthxx) / 2


    comp = funcionario.comp
    comps_adjacentes = competencia_meses_adjacentes(comp)
    beneficios = Beneficios_Mala.objects.filter(codigo_fc=funcionario.codigo_fc, comp__in=comps_adjacentes)


    vt_beneficios = beneficios.filter(tipo_de_beneficio='VT')
    vr_beneficios = beneficios.filter(tipo_de_beneficio='VR/VA')
    cesta_beneficios = beneficios.filter(tipo_de_beneficio='CESTA')



################################      VALE TRANSPORTE    #################################################
    data = [[f"Vale Transporte"], ["Quantidade", "Data de Pagamento", "Valor Pago", "Aut"]]
    for beneficio in vt_beneficios:
        data.append([beneficio.dias_calculados, beneficio.data_de_pagamento, beneficio.valor_pago, beneficio.aut])

    table = Table(data, colWidths=[None, None, None, None])
    table.setStyle(TableStyle([
        # Estilos para o título
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 13),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('SPAN', (0, 0), (-1, 0)),  # Faça o título abranger todas as colunas
        # Estilos gerais
        ('BACKGROUND', (0, 1), (-1, 0), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, 0), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('GRID', (0,1), (-1,-1), 1, colors.black),
        ('BOTTOMPADDING', (0, 1), (-1, 0), 6),  
        ('BOTTOMPADDING', (0, 1), (-1, -1), 0),
        ('LINEABOVE', (0, 2), (-1, 2), 1, colors.black),
    ]))


    table.wrapOn(p, letter[0], letter[1])
    table.drawOn(p, rect_xx, 565)  # Você pode precisar ajustar essas coordenadas para cada tabela.



################################      VALE VR/VA    #################################################


    data = [[f"Vale VR/VA"], ["Quantidade", "Data de Pagamento", "Valor Pago", "Aut"]]
    for beneficio in vr_beneficios:
        data.append([beneficio.dias_calculados, beneficio.data_de_pagamento, beneficio.valor_pago, beneficio.aut])

    table = Table(data, colWidths=[None, None, None, None])
    table.setStyle(TableStyle([
        # Estilos para o título
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 13),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('SPAN', (0, 0), (-1, 0)),  # Faça o título abranger todas as colunas
        # Estilos gerais
        ('BACKGROUND', (0, 1), (-1, 0), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, 0), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('GRID', (0,1), (-1,-1), 1, colors.black),
        ('BOTTOMPADDING', (0, 1), (-1, 0), 6),  
        ('BOTTOMPADDING', (0, 1), (-1, -1), 0),
        ('LINEABOVE', (0, 2), (-1, 2), 1, colors.black),
    ]))

    table.wrapOn(p, letter[0], letter[1])
    table.drawOn(p, rect_xx, 395)  # Você pode precisar ajustar essas coordenadas para cada tabela.





################################      VALE CESTA    #################################################


    data = [[f"Cesta"], ["Quantidade", "Data de Pagamento", "Valor Pago", "Aut"]]
    for beneficio in cesta_beneficios:
        data.append([beneficio.dias_calculados, beneficio.data_de_pagamento, beneficio.valor_pago, beneficio.aut])

    table = Table(data, colWidths=[None, None, None, None])
    table.setStyle(TableStyle([
        # Estilos para o título
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 13),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('SPAN', (0, 0), (-1, 0)),  # Faça o título abranger todas as colunas
        # Estilos gerais
        ('BACKGROUND', (0, 1), (-1, 0), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, 0), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('GRID', (0,1), (-1,-1), 1, colors.black),
        ('BOTTOMPADDING', (0, 1), (-1, 0), 6),  
        ('BOTTOMPADDING', (0, 1), (-1, -1), 0),
        ('LINEABOVE', (0, 2), (-1, 2), 1, colors.black),
    ]))

    table.wrapOn(p, letter[0], letter[1])
    table.drawOn(p, rect_xx, 180)  # Você pode precisar ajustar essas coordenadas para cada tabela.









    # Finalizar a terceira página e começa a 4°.
    p.showPage()
    folha = Folha_de_Ponto.objects.get(codigo_fc=funcionario.codigo_fc, comp=funcionario.comp)
    periodo_apontamento = folha.periodo_apontamento
    # converter a competência para um objeto datetime
    competencia_calculo = datetime.strptime(str(folha.comp), "%Y%m")
    # subtrair um mês da competência
    competencia_calculo_normal = competencia_calculo
    competencia_calculo_menos_um = competencia_calculo - relativedelta(months=1)

    draw_centered_text(p, 805, f"Folha De Ponto", fontsize=16, fontstyle="bold")


    # formatar a competência_menos_um como uma string no formato desejado
    competencia_str = competencia_calculo_menos_um.strftime('%m/%Y')
    competencia_str_normal = competencia_calculo_normal.strftime('%m/%Y')

    dados_dias = []
    if periodo_apontamento == '21 A 20':
        dias_do_mes = list(range(21, 32)) + list(range(1, 21)) 
    elif periodo_apontamento == '01 A 30':
        dias_do_mes = list(range(1, 32))


    # Loop por todos os dias do mês
    for i, dia in enumerate(dias_do_mes):
        # Se dia for maior ou igual a 10, preenche com zero à esquerda
        if dia >= 10:
            dia_str = str(dia).zfill(2)
        else:
            dia_str = str(dia)

    # Determina o mês de acordo com o dia do mês (21 a 31: mês anterior, 1 a 20: mês atual)
        if periodo_apontamento == '21 A 20':
            if dia < 21:
                competencia_str = competencia_str_normal
                competencia_calculo = competencia_calculo_normal
                suffix = '_s'
            else:
                competencia_str = competencia_str
                competencia_calculo = competencia_calculo_menos_um
                suffix = ''
        elif periodo_apontamento == '01 A 30':
            competencia_str = competencia_str_normal
            competencia_calculo = competencia_calculo_normal
            suffix = '_s'


        # Verifica se o dia existe no mês
        if dia > monthrange(competencia_calculo.year, competencia_calculo.month)[1]:
            continue

        # Cria um objeto datetime para o dia atual
        data_atual = datetime(competencia_calculo.year, competencia_calculo.month, dia)

        # Dias da semana: 0 é segunda-feira e 6 é domingo.
        dia_da_semana = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"][data_atual.weekday()]

        # adiciona os dados do dia à lista
        dados_dias.append({
            "data": f"{dia_str}/{competencia_str}", 
            "dia": dia_da_semana, 
            "jornada": getattr(folha, f'{dia_str}_m{suffix}'), 
            "horas": getattr(folha, f'{dia_str}_h{suffix}'), 
            "horas_decimal": getattr(folha, f'{dia_str}_d{suffix}')
        })

    data = [["Data da marcação", "Dia", "Jornada Considerada", "Horas Trabalhadas", "Horas Trabalhadas em Decimal"]]
    for dia in dados_dias:
        data.append([dia["data"], dia["dia"], dia["jornada"], dia["horas"], dia["horas_decimal"]])


    table = Table(data, colWidths=[None, None, 190, None, None])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.white),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),  #Espaçemnto da primeira linha.
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ('LINEABOVE', (0, 1), (-1, 1), 1, colors.black),  # Adicionado
    ]))

    table.wrapOn(p, letter[0], letter[1])
    table.drawOn(p, 10, 310) 








    dados_dias_2 = [
        {"total": "Total HS Trabalhadas:", "horas": f"{folha.total_de_horas}"},
        {"total": "Total HS Normal:", "horas": f"{folha.hs_normais}"},
        {"total": "Total HS Noturnas:", "horas": f"{folha.hs_noturnas}"},
        {"total": "Total HS 50%:", "horas": f"{folha.he_50}"},
        {"total": "Total HS 100%:", "horas": f"{folha.he_100}"},
    
    ]

    data = [["Totais", "Horas"]]
    for dia in dados_dias_2:
        data.append([dia["total"], dia["horas"]])


    table = Table(data, colWidths=[150, 100, 100])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.white),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),  #Espaçemnto da primeira linha.
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ('LINEABOVE', (0, 1), (-1, 1), 1, colors.black),  # Adicionado
    ]))

    table.wrapOn(p, letter[0], letter[1])
    table.drawOn(p, 10, 150)

    p.save()

    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=f'{funcionario.comp} - {funcionario.codigo_fc} - {funcionario.nome}.pdf')

