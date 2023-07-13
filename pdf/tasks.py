from __future__ import absolute_import, unicode_literals

import os

import openpyxl
from celery import shared_task
from django.core.paginator import Paginator
from django.shortcuts import redirect, render

from .models import Beneficios_Mala, Folha_de_Ponto, Funcionario


def non_null_defaults(row, fields):
    return {field: value for field, value in zip(fields, row[1:]) if value is not None}


def non_null_defaults2(row):
    fields = [
        'comp',
        'codigo',
        'codigo_fc',
        'aut',
        'data_inicio',
        'data_fim',
        'dias_calculados',
        'tipo_de_beneficio',
        'valor_pago',
        'data_de_pagamento',
    ]

    return {field: value for field, value in zip(fields, row[1:]) if value is not None}

@shared_task
def importar_excel_beneficios(filepath):
    workbook = openpyxl.load_workbook(filepath, read_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(min_row=2, values_only=True))  # Converter o gerador em uma lista

    paginator = Paginator(rows, 500)  # 1000 linhas por página

    for page_number in paginator.page_range:
        page = paginator.page(page_number)
        for row in page.object_list:
            defaults = non_null_defaults2(row)
            Beneficios_Mala.objects.update_or_create(id=row[0], defaults=defaults)
    
    # Remove the file after processing
    os.remove(filepath)




@shared_task
def importar_excel_funcionario(filepath):
    workbook = openpyxl.load_workbook(filepath, read_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    paginator = Paginator(rows, 50) 

    fields = [
            'comp',
            'cod_cliente',
            'cliente',
            'codigo',
            'codigo_fc',
            'cto',
            'cpf',
            'nome',
            'cargo',
            'adm',
            'dem_compt',
            'valid_desl',
            'tipo',
            'qtde_hs_normais',
            'horas_normais',
            'dissidio_retroativo',
            'dsr_s_horas_normal',
            'qtde_hs_not',
            'adicional_noturno',
            'dsr_s_adicional',
            'qtde_he_50',
            'hora_extra_50',
            'qtde_he_50_not',
            'hora_extra_50_noturno',
            'dsr_s_hora_extra_50',
            'qtde_he_100',
            'hora_extra_100',
            'qtde_he_100_not',
            'hora_extra_100_noturno',
            'dsr_s_hora_extra_100',
            'adic_periculosidade',
            'adicional_de_funcao_25',
            'adicional_de_atividade_30',
            'adicional_final_de_semana_15',
            'salario_familia',
            'falta_abonada_ponto_eletr',
            'qtde_atestado_horistas',
            'atestado_horistas',
            'licenca_remunerada_gestante',
            'salario_maternidade',
            'aux_doenca_15_dias',
            'acidente_de_trabalho_15_dias',
            'acidente_de_trabalho_fgts',
            'fgts_pago_rct',
            'fgts_multa_40',
            'fgts_s_13_salario',
            'fgts_s_aviso_previo',
            'verbas_rescisorias',
            'saldo_negativo_verba_nao_repassada',
            'ferias',
            'um_terco_ferias',
            'decimo_terceiro_salario_indenizado_e_adicionais_considerar',
            'decimo_terceiro_salario_indenizado_e_adicionais',
            'antecip_13_a_vencer',
            'ferias_adiantadas',
            'inss_s_13_salario',
            'inss_s_ferias',
            'inss_s_hora_extra',
            'inss_s_aviso_previo',
            'irrf_s_13_salario',
            'irrf_s_ferias',
            'irrf_s_aviso_previo',
            'inss_s_13_salario_adicional',
            'inss_s_13_salario_abono',
            'inss_s_ferias_vendidas',
            'inss_s_ferias_vendidas_adiantadas',
            'inss_s_ferias_adiantadas',
            'inss_s_13_salario_complementar',
            'irrf_s_ferias_vendidas',
            'irrf_s_ferias_vendidas_adiantadas',
            'irrf_s_ferias_adiantadas',
            'irrf_s_13_salario_complementar',
            'contrib_sindical',
            'contrib_assistencial',
            'contrib_confederativa',
            'contrib_outra',
            'plano_saude',
            'vr',
            'vt',
            'uniforme',
            'seguro_vida',
            'vale_gasolina',
            'vale_cultura',
            'liquido_1',
            'aut_1',
            'liquido_2',
            'aut_2',
            'liquido_3',
            'aut_3',
            'liquido_4',
            'aut_4',
            'liquido_5',
            'aut_5',
            'data_1',
            'data_2',
            'data_3',
            'data_4',
            'data_5',
            'qtde_dsr_feriado',
            'dsr_feriado',
        ] 

    for page_number in paginator.page_range:
        page = paginator.page(page_number)
        for row in page.object_list:
            defaults = non_null_defaults(row, fields)
            Funcionario.objects.update_or_create(id=row[0], defaults=defaults)

    # Remove the file after processing
    os.remove(filepath)


@shared_task
def importar_excel_folha_de_ponto(filepath):
    workbook = openpyxl.load_workbook(filepath, read_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    paginator = Paginator(rows, 50) 

    fields = [
        'comp',
        'codigo',
        'codigo_fc',
        'total_de_horas',
        'hs_normais',
        'hs_noturnas',
        'he_50',
        'he_100',
        '21_m',
        '21_h',
        '21_d',
        '22_m',
        '22_h',
        '22_d',
        '23_m',
        '23_h',
        '23_d',
        '24_m',
        '24_h',
        '24_d',
        '25_m',
        '25_h',
        '25_d',
        '26_m',
        '26_h',
        '26_d',
        '27_m',
        '27_h',
        '27_d',
        '28_m',
        '28_h',
        '28_d',
        '29_m',
        '29_h',
        '29_d',
        '30_m',
        '30_h',
        '30_d',
        '31_m',
        '31_h',
        '31_d',
        '1_m_s',
        '1_h_s',
        '1_d_s',
        '2_m_s',
        '2_h_s',
        '2_d_s',
        '3_m_s',
        '3_h_s',
        '3_d_s',
        '4_m_s',
        '4_h_s',
        '4_d_s',
        '5_m_s',
        '5_h_s',
        '5_d_s',
        '6_m_s',
        '6_h_s',
        '6_d_s',
        '7_m_s',
        '7_h_s',
        '7_d_s',
        '8_m_s',
        '8_h_s',
        '8_d_s',
        '9_m_s',
        '9_h_s',
        '9_d_s',
        '10_m_s',
        '10_h_s',
        '10_d_s',
        '11_m_s',
        '11_h_s',
        '11_d_s',
        '12_m_s',
        '12_h_s',
        '12_d_s',
        '13_m_s',
        '13_h_s',
        '13_d_s',
        '14_m_s',
        '14_h_s',
        '14_d_s',
        '15_m_s',
        '15_h_s',
        '15_d_s',
        '16_m_s',
        '16_h_s',
        '16_d_s',
        '17_m_s',
        '17_h_s',
        '17_d_s',
        '18_m_s',
        '18_h_s',
        '18_d_s',
        '19_m_s',
        '19_h_s',
        '19_d_s',
        '20_m_s',
        '20_h_s',
        '20_d_s',
        '21_m_s',
        '21_h_s',
        '21_d_s',
        '22_m_s',
        '22_h_s',
        '22_d_s',
        '23_m_s',
        '23_h_s',
        '23_d_s',
        '24_m_s',
        '24_h_s',
        '24_d_s',
        '25_m_s',
        '25_h_s',
        '25_d_s',
        '26_m_s',
        '26_h_s',
        '26_d_s',
        '27_m_s',
        '27_h_s',
        '27_d_s',
        '28_m_s',
        '28_h_s',
        '28_d_s',
        '29_m_s',
        '29_h_s',
        '29_d_s',
        '30_m_s',
        '30_h_s',
        '30_d_s',
        '31_m_s',
        '31_h_s',
        '31_d_s',
        'periodo_apontamento',
        ]

    for page_number in paginator.page_range:
        page = paginator.page(page_number)
        for row in page.object_list:
            defaults = non_null_defaults(row, fields)
            Folha_de_Ponto.objects.update_or_create(id=row[0], defaults=defaults)

    # Remove the file after processing
    os.remove(filepath)










