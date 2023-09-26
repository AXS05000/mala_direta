from __future__ import absolute_import, unicode_literals

import os
from datetime import date

import openpyxl
from celery import shared_task
from django.core.paginator import Paginator
from django.shortcuts import render
from zeep import Client

from .models import BaseCNPJ, BaseInfoContratos, NotaFiscal2, Notas


@shared_task
def task_atualizar_notas():
    atualizar_notas_automaticamente()



def non_null_defaults(row, fields):
    return {field: value for field, value in zip(fields, row[1:]) if value is not None}



@shared_task
def import_basecnpj_from_excel(filepath):
    workbook = openpyxl.load_workbook(filepath, read_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    paginator = Paginator(rows, 50) 

    fields = [
            'cnpj',
            'razao',
            'avenida_rua',
            'endereco',
            'numero',
            'complemento',
            'bairro',
            'municipio',
            'uf',
            'cep',
            'nome_cliente',
            'tipo_de_servico',
            'iss',
            'unidade',
            'mcu',
            'tipo_de_cliente',
            'tx_adm',
        ] 

    for page_number in paginator.page_range:
        page = paginator.page(page_number)
        for row in page.object_list:
            defaults = non_null_defaults(row, fields)
            BaseCNPJ.objects.update_or_create(id=row[0], defaults=defaults)

    # Remove the file after processing
    os.remove(filepath)


@shared_task
def import_notas_from_excel(filepath):
    workbook = openpyxl.load_workbook(filepath, read_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    paginator = Paginator(rows, 50)

    # Listando todos os campos
    fields = [
        'data_de_criacao',
        'data_de_modificacao',
        'baseinfocontratos_id',
        'nota_cancelada',
        'competencia_nota_id',
        'tipo_de_faturamento',
        'quantidade_hora',
        'baseinfocontratos2_id',
        'quantidade_hora2',
        'baseinfocontratos3_id',
        'quantidade_hora3',
        'baseinfocontratos4_id',
        'quantidade_hora4',
        'baseinfocontratos5_id',
        'quantidade_hora5',
        'baseinfocontratos6_id',
        'quantidade_hora6',
        'baseinfocontratos7_id',
        'quantidade_hora7',
        'baseinfocontratos8_id',
        'quantidade_hora8',
        'cnpj_da_nota_id',
        'texto_livre',
        'contrato_texto_livre',
        'total_valor_outros',
        'porcentagem_ans',
        'competencia_nota_ans_id'
    ]

    for page_number in paginator.page_range:
        page = paginator.page(page_number)
        for row in page.object_list:
            # Aqui você pode adicionar lógica para lidar com campos especiais
            # Por exemplo, se você tem um ForeignKey para BaseInfoContratos, você pode buscar o objeto assim:
            # baseinfocontratos = BaseInfoContratos.objects.get(id=row[indice_do_campo])
            # E então adicionar ao seu dict defaults:
            # defaults['baseinfocontratos'] = baseinfocontratos
            defaults = non_null_defaults(row, fields)
            Notas.objects.update_or_create(id=row[0], defaults=defaults)

    os.remove(filepath)



@shared_task
def import_baseinfo_from_excel(filepath):
    workbook = openpyxl.load_workbook(filepath, read_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    paginator = Paginator(rows, 50) 

    fields = [
            'cod_cliente',
            'contrato',
            'cargo',
            'valor_hora',
            'data_inicio_cto',
            'contrato_ativo',
        ] 

    for page_number in paginator.page_range:
        page = paginator.page(page_number)
        for row in page.object_list:
            defaults = non_null_defaults(row, fields)
            BaseInfoContratos.objects.update_or_create(id=row[0], defaults=defaults)

    # Remove the file after processing
    os.remove(filepath)







    



#CONSULTA NA API
def consultar_api():
    # Criando o objeto cliente SOAP
    client = Client('https://nfe.osasco.sp.gov.br/EISSNFEWebServices/NotaFiscalEletronica.svc?wsdl')

    # Criando o request com os dados passados
    request_data = {
        'ChaveAutenticacao': '5eb04d8c-fd9a-49ba-ab45-d06d816df7ad',  # valor fixo
        'DataInicial': date(2023, 5, 1),  # valor fixo
        'DataFinal': date.today(),  # data de hoje
        'NumeroReciboInicial': None,  # valor fixo
        'NumeroReciboFinal': None,  # valor fixo
        'NumeroReciboUnico': None  # valor fixo
    }

    # Fazendo a requisição e obtendo a resposta
    response = client.service.Consultar(request=request_data)

    # Retornando a resposta
    return response







# PUXAR NOTAS PARA O MODELS.
def atualizar_notas(request):
    if request.method == 'POST':
        response = consultar_api()

        if not response['Erro']:
            notas_geradas = response['NotasGeradas']['NotaFiscalConsultaDTO']
            NotaFiscal2.objects.all().delete()  # Remove as notas existentes

            for nota in notas_geradas:
                NotaFiscal2.objects.create(
                    aliquota = nota['Aliquota'],
                    cod_atividade = nota['CodAtividade'].strip(),
                    cod_obra = nota['CodObra'],
                    codigo_autenticidade = nota['CodigoAutenticidade'],
                    data_cancelamento = nota['DataCancelamento'],
                    data_emissao = nota['DataEmissao'],
                    data_recibo = nota['DataRecibo'],
                    doc_tomador = nota['DocTomador'],
                    endereco_prestacao_servico = nota['EnderecoPrestacaoServico'],
                    link_nfe = nota['LinkNFE'],
                    motivo_cancelamento = nota['MotivoCancelamento'],
                    nome_tomador = nota['NomeTomador'],
                    nosso_numero = nota['NossoNumero'],
                    numero = nota['Numero'],
                    numero_recibo = nota['NumeroRecibo'],
                    substituicao_tributaria = nota['SubstituicaoTributaria'],
                    valor = nota['Valor'],
                    valor_iss = nota['ValorIss'],
                    valor_nfe = nota['ValorNFE']
                )

            return render(request, 'notas/update.html', {'notas': NotaFiscal2.objects.all()})
        else:
            print(f"Erro: {response['MensagemErro']}")

    return render(request, 'notas/update.html')


def atualizar_notas_automaticamente():
    response = consultar_api()

    if not response['Erro']:
        notas_geradas = response['NotasGeradas']['NotaFiscalConsultaDTO']
        NotaFiscal2.objects.all().delete()  # Remove as notas existentes

        for nota in notas_geradas:
                NotaFiscal2.objects.create(
                    aliquota = nota['Aliquota'],
                    cod_atividade = nota['CodAtividade'].strip(),
                    cod_obra = nota['CodObra'],
                    codigo_autenticidade = nota['CodigoAutenticidade'],
                    data_cancelamento = nota['DataCancelamento'],
                    data_emissao = nota['DataEmissao'],
                    data_recibo = nota['DataRecibo'],
                    doc_tomador = nota['DocTomador'],
                    endereco_prestacao_servico = nota['EnderecoPrestacaoServico'],
                    link_nfe = nota['LinkNFE'],
                    motivo_cancelamento = nota['MotivoCancelamento'],
                    nome_tomador = nota['NomeTomador'],
                    nosso_numero = nota['NossoNumero'],
                    numero = nota['Numero'],
                    numero_recibo = nota['NumeroRecibo'],
                    substituicao_tributaria = nota['SubstituicaoTributaria'],
                    valor = nota['Valor'],
                    valor_iss = nota['ValorIss'],
                    valor_nfe = nota['ValorNFE']
                )

        print("Notas atualizadas com sucesso.")
    else:
        print(f"Erro: {response['MensagemErro']}")